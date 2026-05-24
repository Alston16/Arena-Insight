import os
import json
import threading
from pathlib import Path
from typing import Any
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import Workbook, load_workbook
from tqdm import tqdm
from dotenv import load_dotenv
from query_processor import QueryProcessor
from llm_factory import create_llm


def load_json_list(file_path: str | Path) -> list[dict[str, Any]]:
    path = Path(file_path)

    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("Expected top-level JSON value to be a list.")

    # Optional: ensure each item is a JSON object (Python dict)
    if not all(isinstance(item, dict) for item in data):
        raise ValueError("Expected every item in the list to be a JSON object.")

    return data


def get_or_create_workbook(excel_path: Path):
    if excel_path.exists():
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "results"
        sheet.append(["input", "context", "actual_output", "expected_output", "routed_agent"])
        workbook.save(excel_path)
    return workbook, sheet


items = load_json_list("qafinalclean.json")
print(f"Loaded {len(items)} records")
if not items:
    raise ValueError("No records found in qafinalclean.json")


load_dotenv()

excel_file = Path("deep_test_results.xlsx")
workbook, sheet = get_or_create_workbook(excel_file)

# Resume from the first unprocessed item based on rows already persisted.
processed_count = max(sheet.max_row - 1, 0)
if processed_count >= len(items):
    print("All items are already processed. Nothing to do.")
else:
    print(f"Resuming from item index {processed_count}")

# Thread safe workbook writing
workbook_lock = threading.Lock()
thread_local = threading.local()

def get_thread_query_processor():
    if not hasattr(thread_local, "query_processor"):
        thread_local.query_processor = QueryProcessor(
            create_llm(),
            verbose=False
        )
    return thread_local.query_processor

def process_single_item(item):
    query_input = item.get("input", "")
    expected_output = item.get("output", item.get("expected_output", ""))

    qp = get_thread_query_processor()
    try:
        actual_output = qp.processQuery(query_input, [])
        context = qp.get_generation_context()
        routed_agent = qp.get_route_decision()
    except Exception as exc:
        actual_output = f"ERROR: {exc}"
        context = ""
        routed_agent = ""

    with workbook_lock:
        sheet.append([
            query_input,
            context,
            actual_output,
            expected_output,
            routed_agent,
        ])
        workbook.save(excel_file)

if processed_count < len(items):
    remaining_items = items[processed_count:]
    max_workers = int(os.getenv("MAX_WORKERS", "4"))
    print(f"Starting parallel queue processor with {max_workers} workers...")
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_single_item, item): item for item in remaining_items}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Processing queries", unit="query"):
            # We just consume the iterator to keep tqdm updated
            try:
                future.result()
            except Exception as e:
                print(f"Task generated an exception: {e}")

workbook.close()
print(f"Completed. Results are in {excel_file}")
